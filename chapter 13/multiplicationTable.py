import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
N = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
fontObj1 = Font(bold = True)
for i in range(2, N+2):
    sheet['A'+str(i)] = i-1
    sheet['A'+str(i)].font = fontObj1
    sheet[get_column_letter(i)+str(1)] = i-1
    sheet[get_column_letter(i)+str(1)].font = fontObj1

for rowOfCellObjects in sheet['A2':get_column_letter(N+1)+str(N+1)]:
    mult1 = rowOfCellObjects[0].value
    for cellObj in rowOfCellObjects[1:]:
        mult2 = sheet[get_column_letter(cellObj.column)+str(1)].value
        cellObj.value = mult1 * mult2

wb.save('mult.xlsx')
