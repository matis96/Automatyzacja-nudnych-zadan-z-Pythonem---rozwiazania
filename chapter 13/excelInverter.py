import openpyxl
file = 'inverter.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb.active
wb.create_sheet('temp')
ws_temp = wb['temp']

for rowOfCellObjects in ws:
    for cellObj in rowOfCellObjects:
        ws_temp.cell(row = cellObj.column, column = cellObj.row).value = cellObj.value

title = ws.title
del wb[title]
ws_temp.title = title

wb.save(file)

