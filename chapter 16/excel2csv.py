import os
import openpyxl
import csv

for excelFile in os.listdir('.'):
    if excelFile.endswith('.xlsx'):
        wb = openpyxl.load_workbook(excelFile)
        for sheetName in wb.sheetnames:
            sheet = wb[sheetName]
            csvName = excelFile[:-5] + '_' + sheetName + '.csv'
            csvFile = open(csvName, 'w', newline='')
            outputWriter = csv.writer(csvFile)
            for rowNum in range(1, sheet.max_row+1):
                rowData = []
                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(row=rowNum, column=colNum).value)
                outputWriter.writerow(rowData)
            csvFile.close()
        wb.close()
